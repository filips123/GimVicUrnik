from __future__ import annotations

import datetime
import logging
import os
import re
import typing
from urllib.parse import urlparse

from bs4 import BeautifulSoup, ParserRejectedMarkup
from openpyxl import load_workbook
from sqlalchemy import insert

from .base import BaseMultiUpdater, DocumentInfo
from ..database import DocumentType, LunchMenu, SnackMenu
from ..errors import MenuApiError, MenuDateError, MenuFormatError
from ..utils.pdf import extract_tables
from ..utils.sentry import with_span

if typing.TYPE_CHECKING:
    from typing import Any
    from collections.abc import Iterator
    from io import BytesIO
    from sqlalchemy.orm import Session
    from sentry_sdk.tracing import Span
    from ..config import ConfigSourcesMenu


class MenuUpdater(BaseMultiUpdater):
    source = "website"
    error = MenuApiError

    def __init__(self, config: ConfigSourcesMenu, session: Session):
        self.logger = logging.getLogger(__name__)
        self.config = config
        self.session = session

        super().__init__()

    def get_documents(self) -> Iterator[DocumentInfo]:
        """Download and parse the website to retrieve all menu URLs."""

        try:
            response = self.requests.get(self.config.url)
            response.raise_for_status()

            soup = with_span(op="soup")(BeautifulSoup)(response.text, features="lxml")

        except (OSError, ParserRejectedMarkup) as error:
            raise MenuApiError("Error while downloading or parsing menu index") from error

        menus = soup.find_all("li", {"class": "jedilnik"})

        if not menus:
            self.logger.info("No menus found")
            return iter(())

        for menu in menus:
            for link in menu.find_all("a", href=True):
                contents = str(link.contents[0]).lower()

                if "malica" in contents:
                    menu_type = DocumentType.SNACK_MENU
                elif "kosilo" in contents:
                    menu_type = DocumentType.LUNCH_MENU
                else:
                    continue

                menu_url = self.config.url + link["href"]
                menu_extension = os.path.splitext(urlparse(menu_url).path)[1][1:]
                yield DocumentInfo(type=menu_type, url=menu_url, extension=menu_extension)

    def get_document_title(self, document: DocumentInfo) -> str:
        """Return the normalized document title from its type."""

        if document.type == DocumentType.SNACK_MENU:
            return "Jedilnik za malico"
        elif document.type == DocumentType.LUNCH_MENU:
            return "Jedilnik za kosilo"
        else:
            # This cannot happen because only menus are provided by the API
            raise KeyError("Unknown document type for menu")

    def get_document_effective(self, document: DocumentInfo) -> datetime.date:
        """Return the document effective date in a local timezone from the URL."""

        # jedilnik-kosilo-YYYY-MM-DD(-popravek).pdf
        # jedilnik-malica-YYYY-MM-DD(-popravek).pdf
        date = re.search(
            r"jedilnik-(?:kosilo|malica)-(\d+)-(\d+)-(\d+)(?:-[\w-]*)?\.(?:pdf|xlsx)", document.url
        )

        # The specified date is commonly Monday of the effective week
        # However, in some cases, it may also be another day of that week
        # We need to convert it to Monday, so it is stored correctly

        if date:
            specified = datetime.date(
                year=int(date.group(1)),
                month=int(date.group(2)),
                day=int(date.group(3)),
            )

            effective = specified - datetime.timedelta(days=specified.weekday())
            return effective

        raise MenuDateError("Unknown menu date URL format: " + document.url.rsplit("/", 1)[1])

    def document_needs_parsing(self, document: DocumentInfo) -> bool:
        """Return whether the document needs parsing."""

        # All documents provided by this updater need parsing
        return True

    @with_span(op="parse", pass_span=True)
    def parse_document(self, document: DocumentInfo, stream: BytesIO, effective: datetime.date, span: Span) -> None:  # type: ignore[override]
        """Parse the document and store extracted data."""

        span.set_tag("document.source", self.source)
        span.set_tag("document.type", document.type.value)
        span.set_tag("document.format", document.extension)

        match (document.type, document.extension):
            case (DocumentType.SNACK_MENU, "pdf"):
                self._parse_snack_menu_pdf(stream, effective)
            case (DocumentType.LUNCH_MENU, "pdf"):
                self._parse_lunch_menu_pdf(stream, effective)
            case (DocumentType.SNACK_MENU, "xlsx"):
                self._parse_snack_menu_xlsx(stream, effective)
            case (DocumentType.LUNCH_MENU, "xlsx"):
                self._parse_lunch_menu_xlsx(stream, effective)
            case (DocumentType.SNACK_MENU, _):
                raise MenuFormatError("Unknown snack menu document format: " + str(document.extension))
            case (DocumentType.LUNCH_MENU, _):
                raise MenuFormatError("Unknown lunch menu document format: " + str(document.extension))
            case _:
                raise KeyError("Unknown document type for menu")

    def _parse_snack_menu_pdf(self, stream: BytesIO, effective: datetime.date) -> None:
        """Parse the snack menu PDF document."""

        # Extract all tables from a PDF stream
        tables = with_span(op="extract")(extract_tables)(stream)

        days = 0

        # Parse tables into menus and store them
        for table in tables:
            for row in table:
                if not row[1] or "NV in N" in row[1]:
                    continue

                current = effective + datetime.timedelta(days=days)
                days += 1

                menu = {
                    "date": current,
                    "normal": row[1],
                    "poultry": row[2],
                    "vegetarian": row[3],
                    "fruitvegetable": row[4],
                }

                model = self.session.query(SnackMenu).filter(SnackMenu.date == current).first()

                if not model:
                    model = SnackMenu()

                for key, value in menu.items():
                    setattr(model, key, value)

                self.session.add(model)

    def _parse_snack_menu_xlsx(self, stream: BytesIO, effective: datetime.date) -> None:
        """Parse the snack menu XLSX document."""

        # Extract workbook from an XLSX stream
        wb = with_span(op="extract")(load_workbook)(stream, read_only=True, data_only=True)

        snack_menu: dict[str, Any] = {
            "normal": [],
            "poultry": [],
            "vegetarian": [],
            "fruitvegetable": [],
        }
        days = 0

        # Parse menus and store them
        for ws in wb:
            for wr in ws.iter_rows(min_row=2, max_col=5):
                if days == 5:
                    break

                # Ignore blank cells
                if not wr[1].value:
                    continue

                # Check for correct cell value type (else mypy complains)
                if typing.TYPE_CHECKING:
                    assert isinstance(wr[1].value, str)
                    assert isinstance(wr[2].value, str)
                    assert isinstance(wr[3].value, str)
                    assert isinstance(wr[4].value, str)

                # Ignore information cells
                if "NV in N" in wr[1].value:
                    continue

                if wr[1].value:
                    snack_menu["normal"].append(wr[1].value.strip())

                if wr[2].value:
                    snack_menu["poultry"].append(wr[2].value.strip())

                if wr[3].value:
                    snack_menu["vegetarian"].append(wr[3].value.strip())

                if wr[4].value:
                    snack_menu["fruitvegetable"].append(wr[4].value.strip())

                # Store the menu after the end of day
                if wr[0].border.bottom.color:
                    snack_menu["date"] = effective + datetime.timedelta(days=days)
                    self.session.query(SnackMenu).filter(SnackMenu.date == snack_menu["date"]).delete()

                    snack_menu["normal"] = "\n".join(snack_menu["normal"])
                    snack_menu["poultry"] = "\n".join(snack_menu["poultry"])
                    snack_menu["vegetarian"] = "\n".join(snack_menu["vegetarian"])
                    snack_menu["fruitvegetable"] = "\n".join(snack_menu["fruitvegetable"])

                    self.session.execute(insert(SnackMenu), snack_menu)

                    # Set for next day
                    days += 1
                    snack_menu = {
                        "normal": [],
                        "poultry": [],
                        "vegetarian": [],
                        "fruitvegetable": [],
                    }

        wb.close()

    def _parse_lunch_menu_pdf(self, stream: BytesIO, effective: datetime.date) -> None:
        """Parse the lunch menu PDF document."""

        # Extract all tables from a PDF stream
        tables = with_span(op="extract")(extract_tables)(stream)

        days = 0

        # Parse tables into menus and store them
        for table in tables:
            for row in table:
                if not row[1] or "N KOSILO" in row[1]:
                    continue

                current = effective + datetime.timedelta(days=days)
                days += 1

                menu = {
                    "date": current,
                    "normal": row[1],
                    "vegetarian": row[2],
                }

                model = self.session.query(LunchMenu).filter(LunchMenu.date == current).first()

                if not model:
                    model = LunchMenu()

                for key, value in menu.items():
                    setattr(model, key, value)

                self.session.add(model)

    def _parse_lunch_menu_xlsx(self, stream: BytesIO, effective: datetime.date) -> None:
        """Parse the lunch menu XLSX document."""

        # Extract workbook from an XLSX stream
        wb = with_span(op="extract")(load_workbook)(stream, read_only=True, data_only=True)

        lunch_menu: dict[str, Any] = {
            "normal": [],
            "vegetarian": [],
        }
        days = 0

        # Parse menus and store them
        for ws in wb:
            for wr in ws.iter_rows(min_row=2, max_col=3):
                if days == 5:
                    break

                # Ignore blank cells
                if not wr[1].value:
                    continue

                # Check for correct cell value type (else mypy complains)
                if typing.TYPE_CHECKING:
                    assert isinstance(wr[1].value, str)
                    assert isinstance(wr[2].value, str)

                # Ignore information cells
                if "N KOSILO" in wr[1].value:
                    continue

                if wr[1].value:
                    lunch_menu["normal"].append(wr[1].value.strip())

                if wr[2].value:
                    lunch_menu["vegetarian"].append(wr[2].value.strip())

                # Store the menu after the end of day
                if wr[0].border.bottom.color:
                    lunch_menu["date"] = effective + datetime.timedelta(days=days)
                    self.session.query(LunchMenu).filter(LunchMenu.date == lunch_menu["date"]).delete()

                    lunch_menu["normal"] = "\n".join(lunch_menu["normal"])
                    lunch_menu["vegetarian"] = "\n".join(lunch_menu["vegetarian"])

                    self.session.execute(insert(LunchMenu), lunch_menu)

                    # Set for next day
                    days += 1
                    lunch_menu = {
                        "normal": [],
                        "vegetarian": [],
                    }

        wb.close()

    def document_needs_extraction(self, document: DocumentInfo) -> bool:
        """Return whether the document content needs to be extracted."""

        # Menu documents do not have content
        return False

    def extract_document(self, document: DocumentInfo, stream: BytesIO) -> str | None:
        """Extract the document content and return it as HTML."""

        # Menu documents do not have content
        return None
