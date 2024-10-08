from __future__ import annotations

import enum
import logging
import os
import re
import typing
from datetime import date, datetime, timezone
from itertools import product
from urllib.parse import urlparse

import mammoth  # type: ignore
from openpyxl import load_workbook
from sqlalchemy import insert

from .base import BaseMultiUpdater, DocumentInfo
from ..database import Class, DocumentType, LunchSchedule, Substitution
from ..errors import (
    ClassroomApiError,
    InvalidRecordError,
    InvalidTokenError,
    LunchScheduleFormatError,
    SubstitutionsFormatError,
)
from ..utils.database import get_or_create
from ..utils.normalizers import (
    format_substitution,
    normalize_classroom_name,
    normalize_other_names,
    normalize_subject_name,
    normalize_teacher_name,
)
from ..utils.pdf import extract_tables
from ..utils.sentry import with_span

if typing.TYPE_CHECKING:
    from typing import Any
    from collections.abc import Iterator
    from io import BytesIO
    from mammoth.documents import Image, Hyperlink  # type: ignore
    from sqlalchemy.orm import Session
    from sentry_sdk.tracing import Span
    from ..config import ConfigSourcesEClassroom


class ParserType(enum.Enum):
    SUBSTITUTIONS = "substitutions"
    LESSON_CHANGE = "lesson-change"
    SUBJECT_CHANGE = "subject-change"
    CLASSROOM_CHANGE = "classroom-change"
    MORE_TEACHERS = "more-teachers"
    RESERVATIONS = "reservations"
    UNKNOWN = "unknown"


class EClassroomUpdater(BaseMultiUpdater):
    source = "eclassroom"
    error = ClassroomApiError

    def __init__(
        self,
        config: ConfigSourcesEClassroom,
        session: Session,
        parse_substitutions: bool,
        parse_lunch_schedules: bool,
        extract_circulars: bool,
    ) -> None:
        self.logger = logging.getLogger(__name__)
        self.config = config
        self.session = session

        self.parse_substitutions = parse_substitutions
        self.parse_lunch_schedules = parse_lunch_schedules
        self.extract_circulars = extract_circulars

        super().__init__()

    def get_documents(self) -> Iterator[DocumentInfo]:
        """Get all documents from the e-classroom."""

        self._mark_course_viewed()

        yield from self._get_internal_urls()
        yield from self._get_external_urls()

    def _mark_course_viewed(self) -> None:
        """Mark the course as viewed, so we are not removed for inactivity."""

        params = {
            "moodlewsrestformat": "json",
        }
        data = {
            "courseid": self.config.course,
            "wstoken": self.config.token,
            "wsfunction": "core_course_view_course",
        }

        try:
            response = self.requests.post(self.config.webserviceUrl, params=params, data=data)
            response.raise_for_status()

        except (OSError, ValueError) as error:
            raise ClassroomApiError("Error while accessing e-classroom API") from error

    def _get_internal_urls(self) -> Iterator[DocumentInfo]:
        params = {
            "moodlewsrestformat": "json",
        }
        data = {
            "courseid": self.config.course,
            "wstoken": self.config.token,
            "wsfunction": "core_course_get_contents",
        }

        try:
            response = self.requests.post(self.config.webserviceUrl, params=params, data=data)
            response.raise_for_status()
            contents = response.json()

        except (OSError, ValueError) as error:
            raise ClassroomApiError("Error while accessing e-classroom API") from error

        # Handle API errors
        if "errorcode" in contents:
            if contents["errorcode"] == "invalidtoken":
                raise InvalidTokenError(contents["message"])
            elif contents["errorcode"] == "invalidrecord":
                raise InvalidRecordError(contents["message"])
            else:
                raise ClassroomApiError(contents["message"])

        # Yield every document name, URL and date
        for content in contents:
            for module in content["modules"]:
                if "contents" not in module or len(module["contents"]) == 0:
                    continue

                url = self.normalize_url(module["contents"][0]["fileurl"])

                modified = (
                    datetime.fromtimestamp(module["contents"][0]["timemodified"], tz=timezone.utc)
                    if module["contents"][0]["timemodified"]
                    else None
                )

                created = (
                    datetime.fromtimestamp(module["contents"][0]["timecreated"], tz=timezone.utc)
                    if module["contents"][0]["timecreated"]
                    else modified
                )

                yield DocumentInfo(
                    url=url,
                    type=self._get_document_type(url),
                    title=module["name"],
                    created=created,
                    modified=modified,
                    extension=os.path.splitext(urlparse(url).path)[1][1:],
                )

    def _get_external_urls(self) -> Iterator[DocumentInfo]:
        params = {
            "moodlewsrestformat": "json",
        }
        data = {
            "wstoken": self.config.token,
            "wsfunction": "mod_url_get_urls_by_courses",
        }

        try:
            response = self.requests.post(self.config.webserviceUrl, params=params, data=data)
            response.raise_for_status()
            contents = response.json()

        except (OSError, ValueError) as error:
            raise ClassroomApiError("Error while accessing e-classroom API") from error

        # Handle API errors
        if "errorcode" in contents:
            if contents["errorcode"] == "invalidtoken":
                raise InvalidTokenError(contents["message"])
            else:
                raise ClassroomApiError(contents["message"])

        # Yield every external URL name, URL and date
        for content in contents["urls"]:
            if content["course"] != self.config.course:
                continue

            url = self.normalize_url(content["externalurl"])

            modified = (
                datetime.fromtimestamp(content["timemodified"], tz=timezone.utc)
                if content["timemodified"]
                else None
            )

            yield DocumentInfo(
                url=url,
                type=self._get_document_type(url),
                title=content["name"],
                created=modified,
                modified=modified,
                extension=os.path.splitext(urlparse(url).path)[1][1:],
            )

    @staticmethod
    def _get_document_type(url: str) -> DocumentType:
        """Get a document type based on its URL."""

        if "www.dropbox.com" in url:
            return DocumentType.SUBSTITUTIONS
        elif "delitev-kosila" in url or "delitevKosila" in url:
            return DocumentType.LUNCH_SCHEDULE
        elif "okroznica" in url.lower() or "okrožnica" in url.lower():
            return DocumentType.CIRCULAR
        else:
            return DocumentType.OTHER

    def normalize_url(self, url: str) -> str:
        """Convert an e-classroom webservice URL to a normal URL."""

        return (
            url.replace(self.config.pluginFileWebserviceUrl, self.config.pluginFileNormalUrl)
            .replace("?forcedownload=1", "")
            .replace("?dl=0", "?raw=1")
            .replace("&dl=0", "&raw=1")
            .replace("?rlkey=", "?raw=1&rlkey=")
        )

    def tokenize_url(self, url: str) -> str:
        """Convert a normal e-classroom URL to a webservice URL and add the token."""

        if self.config.pluginFileNormalUrl in url:
            return (
                url.replace(self.config.pluginFileNormalUrl, self.config.pluginFileWebserviceUrl)
                + "?token="
                + self.config.token
            )

        return url

    def get_document_title(self, document: DocumentInfo) -> str:
        """Return the normalized document title."""

        if document.type == DocumentType.SUBSTITUTIONS:
            return "Nadomeščanja in obvestila"

        if document.type == DocumentType.LUNCH_SCHEDULE:
            return "Razpored kosila"

        assert document.title
        return document.title

    @typing.no_type_check  # Ignored because if regex fails, we cannot do anything
    def get_document_effective(self, document: DocumentInfo) -> date | None:
        """Return the document effective date in a local timezone."""

        if document.type == DocumentType.SUBSTITUTIONS:
            effective = re.search(r"_obvestila_(.+).pdf", document.url, re.IGNORECASE).group(1)
            return datetime.strptime(effective, "%d._%m._%Y").date()

        if document.type == DocumentType.LUNCH_SCHEDULE:
            title = document.title.split(",")[-1].split("-")[-1].strip()
            search = re.search(r"(\d+) *\. *(\d+) *\. *(\d+)", title)
            return date(year=int(search.group(3)), month=int(search.group(2)), day=int(search.group(1)))

    def document_needs_parsing(self, document: DocumentInfo) -> bool:
        """Return whether the document needs parsing."""

        if document.type == DocumentType.SUBSTITUTIONS:
            return self.parse_substitutions

        if document.type == DocumentType.LUNCH_SCHEDULE:
            return self.parse_lunch_schedules

        return False

    @with_span(op="parse", pass_span=True)
    def parse_document(  # type: ignore[override]
        self,
        document: DocumentInfo,
        stream: BytesIO,
        effective: date,
        span: Span,
    ) -> None:
        """Parse the document and store extracted data."""

        span.set_tag("document.source", self.source)
        span.set_tag("document.type", document.type.value)
        span.set_tag("document.format", document.extension)

        match (document.type, document.extension):
            case (DocumentType.SUBSTITUTIONS, "pdf"):
                self._parse_substitutions_pdf(stream, effective)
            case (DocumentType.LUNCH_SCHEDULE, "xlsx"):
                self._parse_lunch_schedule_xlsx(stream, effective)
            case (DocumentType.SUBSTITUTIONS, _):
                raise SubstitutionsFormatError(
                    "Unknown substitutions document format: " + str(document.extension)
                )
            case (DocumentType.LUNCH_SCHEDULE, _):
                raise LunchScheduleFormatError(
                    "Unknown lunch schedule document format: " + str(document.extension)
                )
            case _:
                # This cannot happen because only these types are provided by the API
                raise KeyError("Unknown parsable document type from the e-classroom")

    def document_needs_extraction(self, document: DocumentInfo) -> bool:
        """Return whether the document content needs to be extracted."""

        # Only DOCX documents (circulars and some other) can have content extracted
        if document.extension == "docx":
            return self.extract_circulars

        return False

    @with_span(op="content", pass_span=True)
    def extract_document(self, document: DocumentInfo, content: bytes, span: Span) -> str | None:  # type: ignore[override]
        """Extract the document content and return it as HTML."""

        span.set_tag("document.source", self.source)
        span.set_tag("document.type", document.type.value)
        span.set_tag("document.format", document.extension)

        def ignore_images(_image: Image) -> dict:
            return {}

        def transform_hyperlinks(hyperlink: Hyperlink) -> Hyperlink:
            hyperlink.target_frame = "_blank"
            return hyperlink

        # Convert DOCX to HTML
        result = mammoth.convert_to_html(
            content,
            convert_image=ignore_images,
            transform_document=mammoth.transforms.element_of_type(
                mammoth.documents.Hyperlink,
                transform_hyperlinks,
            ),
        )
        return typing.cast(str, result.value)

    def _parse_substitutions_pdf(self, stream: BytesIO, effective: date) -> None:
        """Parse the substitutions pdf document."""

        # fmt: off
        header_substitutions = ["ODSOTNI UČITELJ/ICA", "URA", "RAZRED", "UČILNICA", "NADOMEŠČA", "PREDMET", "OPOMBA"]
        header_lesson_change = ["RAZRED", "URA", "UČITELJ/ICA", "PREDMETA", "UČILNICA", "OPOMBA"]
        header_subject_change = ["RAZRED", "URA", "UČITELJ", "PREDMET", "UČILNICA", "OPOMBA"]
        header_classroom_change = ["RAZRED", "URA", "UČITELJ/ICA", "PREDMET", "IZ UČILNICE", "V UČILNICO", "OPOMBA"]
        header_more_teachers = ["URA", "UČITELJ", "RAZRED", "UČILNICA", "OPOMBA"]
        header_reservations = ["URA", "UČILNICA", "REZERVIRAL/A", "OPOMBA"]
        # fmt: on

        day = effective.isoweekday()
        substitutions = []

        parser_type = None
        last_original_teacher = None

        # Extract all tables from a PDF stream
        tables = with_span(op="extract")(extract_tables)(stream)

        # Parse tables into substitutions
        for table in tables:
            for row0 in table:
                # We use different variable name here, otherwise mypy complains
                row = [column.replace("\n", " ").strip() if column else "" for column in row0]

                # Get parser type
                if row == header_substitutions:
                    parser_type = ParserType.SUBSTITUTIONS
                    continue
                elif row == header_lesson_change:
                    parser_type = ParserType.LESSON_CHANGE
                    continue
                elif row == header_subject_change:
                    parser_type = ParserType.SUBJECT_CHANGE
                    continue
                elif row == header_classroom_change:
                    parser_type = ParserType.CLASSROOM_CHANGE
                    continue
                elif row == header_more_teachers:
                    parser_type = ParserType.MORE_TEACHERS
                    continue
                elif row == header_reservations:
                    parser_type = ParserType.RESERVATIONS
                    continue
                elif (
                    "Oddelek" in row[0]
                    or "Razred" in row[0]
                    or "dijaki" in row[0]
                    or "RAZREDNIK" in row[1]
                    or "Spoznavanje" in row[1]
                ):
                    parser_type = ParserType.UNKNOWN
                    continue

                # Skip empty rows
                if not any(row) or not row[1]:
                    continue

                # Parse substitutions
                if parser_type == ParserType.SUBSTITUTIONS:
                    # Get basic substitution properties
                    time = int(row[1][:-1]) if row[1] != "PU" else 0
                    subject = normalize_subject_name(row[5])
                    notes = normalize_other_names(row[6])

                    # Get the original teacher if it is specified
                    # Otherwise, use the last specified original teacher
                    # fmt: off
                    original_teacher = normalize_teacher_name(row[0]) if row[0] else last_original_teacher
                    last_original_teacher = original_teacher
                    # fmt: on

                    # Get the new teacher
                    teacher = normalize_teacher_name(row[4])

                    # Get the classroom (which stays the same)
                    # There may be multiple classrooms per row
                    classrooms = [normalize_classroom_name(name) for name in row[3].split(", ")]

                    # Handle multiple classes
                    classes = row[2].replace(". ", "").split(" - ")
                    classes = classes[:-1] if len(classes) > 1 else classes

                    # fmt: off
                    for class_, classroom in product(classes, classrooms):
                        substitutions.append(format_substitution(
                            self.session,
                            effective, day, time,
                            subject, notes,
                            original_teacher, classroom,
                            class_, teacher, classroom,
                        ))
                    # fmt: on

                elif parser_type == ParserType.LESSON_CHANGE:
                    # Get basic substitution properties
                    time = int(row[1][:-1]) if row[1] != "PU" else 0
                    subject = normalize_subject_name(row[3].split(" → ")[1])
                    notes = normalize_other_names(row[5])

                    # Get the original and the new teacher
                    original_teacher = normalize_teacher_name(row[2].split(" → ")[0])
                    teacher = normalize_teacher_name(row[2].split(" → ")[1])

                    # Get the original and the new classrooms
                    # They are commonly the same, but not always
                    # There may also be multiple classrooms per row
                    split_classrooms = row[4].split(" → ")
                    original_classrooms = []
                    classrooms = []

                    if len(split_classrooms) == 1:
                        # Classroom has stayed the same
                        # fmt: off
                        original_classrooms = [normalize_classroom_name(name) for name in split_classrooms[0].split(", ")]
                        classrooms = original_classrooms
                        # fmt: on

                    elif len(split_classrooms) == 2:
                        # Classroom has changed
                        # fmt: off
                        original_classrooms = [normalize_classroom_name(name) for name in split_classrooms[0].split(", ")]
                        classrooms = [normalize_classroom_name(name) for name in split_classrooms[1].split(", ")]
                        # fmt: on

                    # Handle multiple classes
                    classes = row[0].replace(". ", "").split(" - ")
                    classes = classes[:-1] if len(classes) > 1 else classes

                    # fmt: off
                    for class_, original_classroom, classroom in product(classes, original_classrooms, classrooms):
                        substitutions.append(format_substitution(
                            self.session,
                            effective, day, time,
                            subject, notes,
                            original_teacher, original_classroom,
                            class_, teacher, classroom,
                        ))
                    # fmt: on

                elif parser_type == ParserType.SUBJECT_CHANGE:
                    # Get basic substitution properties
                    time = int(row[1][:-1]) if row[1] != "PU" else 0
                    subject = normalize_subject_name(row[3].split(" → ")[1])
                    notes = normalize_other_names(row[5])

                    # Get the teacher (which stays the same)
                    original_teacher = normalize_teacher_name(row[2])
                    teacher = original_teacher

                    # Get the classroom (which stays the same)
                    original_classroom = normalize_classroom_name(row[4])
                    classroom = original_classroom

                    # Handle multiple classes
                    classes = row[0].replace(". ", "").split(" - ")
                    classes = classes[:-1] if len(classes) > 1 else classes

                    # fmt: off
                    for class_ in classes:
                        substitutions.append(format_substitution(
                            self.session,
                            effective, day, time,
                            subject, notes,
                            original_teacher, original_classroom,
                            class_, teacher, classroom,
                        ))
                    # fmt: on

                elif parser_type == ParserType.CLASSROOM_CHANGE:
                    # Get basic substitution properties
                    time = int(row[1][:-1]) if row[1] != "PU" else 0
                    subject = normalize_subject_name(row[3])
                    notes = normalize_other_names(row[6])

                    # Get the teacher (which stays the same)
                    original_teacher = normalize_teacher_name(row[2])
                    teacher = original_teacher

                    # Get the original and the new classrooms
                    # fmt: off
                    original_classrooms = [normalize_classroom_name(name) for name in row[4].split(", ")]
                    classrooms = [normalize_classroom_name(name) for name in row[5].split(", ")]
                    # fmt: on

                    # Handle multiple classes
                    classes = row[0].replace(". ", "").split(" - ")
                    classes = classes[:-1] if len(classes) > 1 else classes

                    # fmt: off
                    for class_, original_classroom, classroom in product(classes, original_classrooms, classrooms):
                        if original_classroom == classroom:
                            continue

                        substitutions.append(format_substitution(
                            self.session,
                            effective, day, time,
                            subject, notes,
                            original_teacher, original_classroom,
                            class_, teacher, classroom,
                        ))
                    # fmt: on

        # Deduplicate substitutions
        substitutions = list({frozenset(subs.items()): subs for subs in substitutions}.values())

        # Remove old substitutions from a database
        self.session.query(Substitution).filter(Substitution.date == effective).delete()

        # Store new substitutions to a database
        if substitutions:
            self.session.execute(insert(Substitution), substitutions)

    def _parse_lunch_schedule_xlsx(self, stream: BytesIO, effective: date) -> None:
        """
        Parse the lunch schedule xlsx document.

        Columns:
        - Time (Ura)
        - Notes (Opombe/Prilagoditev)
        - Class (Razred)
        - Number of students (Stevilo dijakov)
        - Location (Prostor)
        """

        # Extract workbook from an XLSX stream
        wb = with_span(op="extract")(load_workbook)(stream, data_only=True)

        lunch_schedule = []

        # Parse lunch schedule
        for ws in wb:
            for column in ws.column_dimensions.values():
                if column.hidden:
                    ws.delete_cols(1)

            for wr in ws.iter_rows(min_row=3, max_col=5):
                # Check for correct cell value type
                if typing.TYPE_CHECKING:
                    assert isinstance(wr[0].value, datetime)  # Time
                    assert isinstance(wr[1].value, str)  # Notes
                    assert isinstance(wr[2].value, str)  # Class
                    assert isinstance(wr[4].value, str)  # Location

                # Ignore empty and header rows
                if not wr[2].value or "raz" in wr[2].value:
                    continue

                schedule = {
                    "date": effective,
                    "time": wr[0].value if wr[0].value else None,
                    "notes": wr[1].value.strip() if wr[1].value else None,
                    "location": wr[4].value.strip() if wr[4].value else None,
                    "class_id": get_or_create(self.session, model=Class, name=wr[2].value.strip())[0].id
                    if wr[2].value
                    else None,
                }

                lunch_schedule.append(schedule)

        wb.close()

        # Store schedule to a database
        self.session.query(LunchSchedule).filter(LunchSchedule.date == effective).delete()
        self.session.execute(insert(LunchSchedule), lunch_schedule)
